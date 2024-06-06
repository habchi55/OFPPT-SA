import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from tkcalendar import Calendar
import subprocess
import openpyxl
from openpyxl.styles import Alignment
import sqlite3
from tkinter import PhotoImage
import os

# Custom dialog for input with logo
class CustomDialog(simpledialog.Dialog):
    def __init__(self, parent, title, prompt, logo_image):
        self.prompt = prompt
        self.logo_image = logo_image
        super().__init__(parent, title)

    def body(self, master):
        self.iconphoto(False, self.logo_image)
        tk.Label(master, text=self.prompt).grid(row=0, column=0, padx=5, pady=5)
        self.entry = tk.Entry(master)
        self.entry.grid(row=1, column=0, padx=5, pady=5)
        return self.entry

    def apply(self):
        self.result = self.entry.get()

def custom_askstring(title, prompt, logo_image):
    dialog = CustomDialog(root, title, prompt, logo_image)
    return dialog.result

def create_database():
    conn = sqlite3.connect('../users.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            post TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            department TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def add_user(name, post, email, password, department):
    conn = sqlite3.connect('../users.db')
    c = conn.cursor()
    try:
        c.execute('INSERT INTO users (name, post, email, password, department) VALUES (?, ?, ?, ?, ?)',
                  (name, post, email, password, department))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def validate_user(email, password):
    conn = sqlite3.connect('../users.db')
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE email=? AND password=?', (email, password))
    user = c.fetchone()
    conn.close()
    return user

def get_user_details(user_id):
    conn = sqlite3.connect('../users.db')
    c = conn.cursor()
    c.execute('SELECT name, post, email FROM users WHERE id=?', (user_id,))
    user = c.fetchone()
    conn.close()
    if user:
        return {'name': user[0], 'post': user[1], 'contact': user[2]}
    return None

def get_next_reference_number():
    ref_file = "reference_number.txt"

    if not os.path.exists(ref_file):
        with open(ref_file, "w") as file:
            file.write("0")

    with open(ref_file, "r") as file:
        current_ref = int(file.read().strip())

    next_ref = current_ref + 1

    with open(ref_file, "w") as file:
        file.write(str(next_ref))

    return f"DR:SCQ:{next_ref}"

# Store the selected theme globally
selected_theme = "forest-dark"

def apply_theme(window, theme=None):
    if theme is None:
        theme = selected_theme
    style = ttk.Style(window)
    current_dir = os.path.dirname(os.path.abspath(__file__))

    light_theme_path = os.path.join(current_dir, "forest-light.tcl")
    dark_theme_path = os.path.join(current_dir, "forest-dark.tcl")

    if "forest-light" not in style.theme_names():
        window.tk.call("source", light_theme_path)
    if "forest-dark" not in style.theme_names():
        window.tk.call("source", dark_theme_path)
    style.theme_use(theme)

def show_login_window():
    global email_entry, password_entry, login_window, theme_combobox
    login_window = tk.Toplevel()
    login_window.title("Se connecter")

    left_logo_image = tk.PhotoImage(file=r'C:\Users\Nasser\Downloads\logoprogrammenobg.png')
    login_window.left_logo_image = left_logo_image
    login_window.iconphoto(False, left_logo_image)

    frame = ttk.Frame(login_window)
    frame.pack(padx=10, pady=10)

    left_logo_label = tk.Label(frame, image=left_logo_image)
    left_logo_label.grid(row=0, column=0, rowspan=5, padx=10, pady=10)

    tk.Label(frame, text="Email:").grid(row=0, column=1, sticky="w")
    email_entry = tk.Entry(frame)
    email_entry.grid(row=1, column=1, padx=0, pady=0, sticky="ew")

    tk.Label(frame, text="Mot de passe:").grid(row=2, column=1, sticky="w")
    password_entry = tk.Entry(frame, show="*")
    password_entry.grid(row=3, column=1, padx=0, pady=0, sticky="ew")

    tk.Button(frame, text="Se connecter", command=login).grid(row=4, column=1, pady=5, sticky="ew")
    tk.Button(frame, text="Créer un compte", command=create_account).grid(row=5, column=1, pady=5, sticky="ew")

    theme_combobox = ttk.Combobox(frame, values=["forest-dark", "forest-light"], state="readonly")
    theme_combobox.set("Sélectionner le thème")
    theme_combobox.grid(row=6, column=0, pady=5,padx=5, sticky="ew")
    theme_combobox.bind("<<ComboboxSelected>>", change_theme)

    frame.columnconfigure(1, weight=1)

def change_theme(event):
    global selected_theme
    selected_theme = theme_combobox.get()

def create_account():
    def submit():
        name = name_entry.get()
        post = post_entry.get()
        email = email_entry.get()
        password = password_entry.get()
        department = department_combobox.get()

        if name and post and email and password and department:
            if add_user(name, post, email, password, department):
                messagebox.showinfo("Succès", "Compte créé avec succès!")
                create_window.destroy()
            else:
                messagebox.showerror("Erreur", "L'email existe déjà!")
        else:
            messagebox.showwarning("Entrée requise", "Veuillez remplir tous les champs.")

    create_window = tk.Toplevel()
    create_window.title("Créer un compte")
    apply_theme(create_window)

    left_logo_image = tk.PhotoImage(file=r'C:\Users\Nasser\Downloads\Logoofpptnobg.png')
    create_window.left_logo_image = left_logo_image
    create_window.iconphoto(False, left_logo_image)

    frame = ttk.Frame(create_window)
    frame.pack(padx=10, pady=10)

    left_logo_label = tk.Label(frame, image=left_logo_image)
    left_logo_label.grid(row=0, column=0, rowspan=8, padx=10, pady=10)

    tk.Label(frame, text="Nom:").grid(row=0, column=1, sticky="w")
    name_entry = tk.Entry(frame)
    name_entry.grid(row=1, column=1, padx=0, pady=0, sticky="ew")

    tk.Label(frame, text="Poste:").grid(row=2, column=1, sticky="w")
    post_entry = tk.Entry(frame)
    post_entry.grid(row=3, column=1, padx=0, pady=0, sticky="ew")

    tk.Label(frame, text="Email:").grid(row=4, column=1, sticky="w")
    email_entry = tk.Entry(frame)
    email_entry.grid(row=5, column=1, padx=0, pady=0, sticky="ew")

    tk.Label(frame, text="Mot de passe:").grid(row=6, column=1, sticky="w")
    password_entry = tk.Entry(frame, show="*")
    password_entry.grid(row=7, column=1, padx=0, pady=0, sticky="ew")

    tk.Label(frame, text="Département:").grid(row=8, column=1, sticky="w")
    department_combobox = ttk.Combobox(frame, values=["Direction Régionale", "Complexe de Formation Professionnel"],
                                       state="readonly")
    department_combobox.grid(row=9, column=1, padx=0, pady=0, sticky="ew")

    tk.Button(frame, text="Soumettre", command=submit).grid(row=10, column=1, pady=5, sticky="ew")

    frame.columnconfigure(1, weight=1)

def login():
    global current_user, email_button, selected_theme
    email = email_entry.get()
    password = password_entry.get()

    user = validate_user(email, password)
    if user:
        messagebox.showinfo("Succès", f"Bienvenue {user[1]}!")
        current_user = user
        login_window.destroy()
        root.deiconify()
        apply_theme(root, selected_theme)  # Apply the stored theme after login
        initialize_main_application()

        if current_user[5] != "Direction Régionale":
            email_button.config(state=tk.DISABLED)
    else:
        messagebox.showerror("Erreur", "Email ou mot de passe invalide!")

def imported_file():
    global imported_file_path
    imported_file_path = filedialog.askopenfilename(title="Importer des ressources",
                                                    filetypes=(("PDF files", "*.pdf"), ("All files", "*.*")))
    if imported_file_path:
        print("Ressource sélectionnée:", imported_file_path)

def open_file_from_treeview(event):
    item = treeview.selection()[0]
    file_path = treeview.item(item, "values")[-1]
    if file_path:
        try:
            subprocess.Popen(['start', '', file_path], shell=True)
        except Exception as e:
            print(f"Erreur d'ouverture du fichier: {e}")

def generate_email():
    logo_image = tk.PhotoImage(file=r'C:\Users\Nasser\Downloads\Logoofpptnobg.png')

    complex_name = custom_askstring("Nom du complexe", "Entrez le nom du complexe :", logo_image)

    user_details = get_user_details(current_user[0])
    reference_number = get_next_reference_number()

    if complex_name and user_details:
        email_body = (
            f"Objet : Demande de plans d'action\n"
            f"Référence : {reference_number}\n\n"
            f"Madame, Monsieur,\n\n"
            f"Nous vous prions de bien vouloir nous transmettre les plans d'action couvrant tous les EFPs sous votre gestion pour le complexe {complex_name}. "
            f"Veuillez vous assurer que ces plans respectent les champs du plateforme.\n\n"
            f"Merci de votre collaboration.\n\n"
            f"Cordialement,\n\n"
            f"{user_details['name']}\n"
            f"{user_details['post']}\n"
            f"{user_details['contact']}"
        )

        email_window = tk.Toplevel(root)
        email_window.title("Email")
        email_window.iconphoto(False, logo_image)

        email_text = tk.Text(email_window, wrap='word', height=20, width=80)
        email_text.insert('1.0', email_body)
        email_text.pack(pady=10)

        copy_button = ttk.Button(email_window, text="Copier dans le presse-papiers",
                                 command=lambda: copy_to_clipboard(email_body))
        copy_button.pack(pady=5)
    else:
        messagebox.showwarning("Champs requis",
                               "Veuillez fournir le nom du complexe et vous assurer que les détails de l'utilisateur sont disponibles.")

def copy_to_clipboard(text):
    root.clipboard_clear()
    root.clipboard_append(text)
    messagebox.showinfo("Copié", "Email copié dans le presse-papiers !")

# Define the new column names in the desired order
new_cols = ["CF", "EFP", "VOLET", "ACTION", "OBJET", "CFP/ EFP", "PRIORITÉ", "STATUT", "DEADLINE", "OBSERVATIONS", "RESSOURCES"]

def load_data():
    path = r"C:\Users\Nasser\Documents\test.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)

    # Set up Treeview columns and headings, starting from "Action"
    for col_index, col_name in enumerate(new_cols[3:]):  # Skip the first 3 columns
        treeview.heading(col_index, text=col_name)
        treeview.column(col_index, width=200)

    # Insert rows into the Treeview, starting from the 4th column
    for value_tuple in list_values[1:]:
        value_tuple = [" " if value is None else value for value in value_tuple]
        display_values = value_tuple[3:]  # Skip the first 3 columns
        treeview.insert('', tk.END, values=display_values)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    treeview.bind("<Double-1>", open_file_from_treeview)

# Assuming the Treeview widget is created somewhere in your GUI setup
cols = new_cols[3:]  # Only include columns starting from "ACTION"

def delete_row():
    selected_items = treeview.selection()
    if selected_items:
        path = r"C:\Users\Nasser\Documents\test.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        for item in selected_items:
            values = treeview.item(item)['values']
            treeview.delete(item)
            for row in sheet.iter_rows():
                row_values = [cell.value for cell in row]
                if all(value == selected_value or selected_value == ' ' for value, selected_value in
                       zip(row_values[3:], values) if selected_value.strip()):  # Skip the first 3 columns
                    sheet.delete_rows(row[0].row)
                    break
        workbook.save(path)
    else:
        messagebox.showerror("Erreur", "Veuillez sélectionner au moins une action.")

def insert_row():
    cf = cf_combobox.get()
    efp = efp_combobox.get()
    vol = vol_combobox.get()
    action = action_entry.get()
    objet = objet_entry.get()
    cf_efp = cf_efp_entry.get()
    priorité = priorité_spinbox.get()
    statut = statut_combobox.get()
    deadline = deadline_entry.get()
    observations = observation_entry.get()
    ressources = imported_file_path if imported_file_path else ""

    row_values = [cf, efp, vol, action, objet, cf_efp, priorité, statut, deadline, observations, ressources]

    treeview.insert('', tk.END, values=row_values[3:])  # Insert values starting from the 4th column

    path = r"C:\Users\Nasser\Documents\test.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.append(row_values)
    workbook.save(path)

def manage_database():
    password = custom_askstring("Mot de passe", "Entrez le mot de passe admin:", root.logo)
    if password == "scq2024":
        show_user_management_window()
    else:
        messagebox.showerror("Erreur", "Mot de passe invalide!")

def show_user_management_window():
    user_window = tk.Toplevel(root)
    user_window.title("Gérer les utilisateurs")
    apply_theme(user_window)
    user_window.iconphoto(False, root.logo)

    frame = ttk.Frame(user_window)
    frame.pack(padx=10, pady=10, fill='both', expand=True)

    user_tree = ttk.Treeview(frame, columns=("ID", "Nom", "Poste", "Email", "Département"), show='headings')
    user_tree.heading("ID", text="ID")
    user_tree.heading("Nom", text="Nom")
    user_tree.heading("Poste", text="Poste")
    user_tree.heading("Email", text="Email")
    user_tree.heading("Département", text="Département")
    user_tree.pack(fill='both', expand=True)

    def load_users():
        conn = sqlite3.connect('../users.db')
        c = conn.cursor()
        c.execute('SELECT id, name, post, email, department FROM users')
        users = c.fetchall()
        conn.close()
        for user in users:
            user_tree.insert('', 'end', values=user)

    def add_user_window():
        def add_user_to_db():
            name = name_entry.get()
            post = post_entry.get()
            email = email_entry.get()
            password = password_entry.get()
            department = department_combobox.get()

            if name and post and email and password and department:
                if add_user(name, post, email, password, department):
                    messagebox.showinfo("Succès", "Utilisateur ajouté avec succès!")
                    user_tree.insert('', 'end',
                                     values=(user_tree.get_children()[-1] + 1, name, post, email, department))
                    add_user_win.destroy()
                else:
                    messagebox.showerror("Erreur", "L'email existe déjà!")
            else:
                messagebox.showwarning("Entrée requise", "Veuillez remplir tous les champs.")

        add_user_win = tk.Toplevel(user_window)
        add_user_win.title("Ajouter un utilisateur")
        apply_theme(add_user_win)
        add_user_win.iconphoto(False, root.logo)

        frame = ttk.Frame(add_user_win)
        frame.pack(padx=10, pady=10)

        tk.Label(frame, text="Nom:").grid(row=0, column=0, sticky="w")
        name_entry = tk.Entry(frame)
        name_entry.grid(row=1, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Poste:").grid(row=2, column=0, sticky="w")
        post_entry = tk.Entry(frame)
        post_entry.grid(row=3, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Email:").grid(row=4, column=0, sticky="w")
        email_entry = tk.Entry(frame)
        email_entry.grid(row=5, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Mot de passe:").grid(row=6, column=0, sticky="w")
        password_entry = tk.Entry(frame, show="*")
        password_entry.grid(row=7, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Département:").grid(row=8, column=0, sticky="w")
        department_combobox = ttk.Combobox(frame, values=["Direction Régionale", "Complexe de Formation Professionnel"],
                                           state="readonly")
        department_combobox.grid(row=9, column=0, padx=0, pady=0, sticky="ew")

        tk.Button(frame, text="Ajouter l'utilisateur", command=add_user_to_db).grid(row=10, column=0, pady=5, sticky="ew")

    def delete_user():
        selected_item = user_tree.selection()[0]
        user_id = user_tree.item(selected_item, "values")[0]
        conn = sqlite3.connect('../users.db')
        c = conn.cursor()
        c.execute('DELETE FROM users WHERE id=?', (user_id,))
        conn.commit()
        conn.close()
        user_tree.delete(selected_item)
        messagebox.showinfo("Succès", "Utilisateur supprimé avec succès!")

    def update_user_window():
        selected_item = user_tree.selection()[0]
        user = user_tree.item(selected_item, "values")

        def update_user_in_db():
            name = name_entry.get()
            post = post_entry.get()
            email = email_entry.get()
            password = password_entry.get()
            department = department_combobox.get()

            if name and post and email and password and department:
                conn = sqlite3.connect('../users.db')
                c = conn.cursor()
                c.execute('UPDATE users SET name=?, post=?, email=?, password=?, department=? WHERE id=?',
                          (name, post, email, password, department, user[0]))
                conn.commit()
                conn.close()
                user_tree.item(selected_item, values=(user[0], name, post, email, department))
                messagebox.showinfo("Succès", "Utilisateur mis à jour avec succès!")
                update_user_win.destroy()
            else:
                messagebox.showwarning("Entrée requise", "Veuillez remplir tous les champs.")

        update_user_win = tk.Toplevel(user_window)
        update_user_win.title("Mettre à jour l'utilisateur")
        apply_theme(update_user_win)
        update_user_win.iconphoto(False, root.logo)

        frame = ttk.Frame(update_user_win)
        frame.pack(padx=10, pady=10)

        tk.Label(frame, text="Nom:").grid(row=0, column=0, sticky="w")
        name_entry = tk.Entry(frame)
        name_entry.insert(0, user[1])
        name_entry.grid(row=1, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Poste:").grid(row=2, column=0, sticky="w")
        post_entry = tk.Entry(frame)
        post_entry.insert(0, user[2])
        post_entry.grid(row=3, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Email:").grid(row=4, column=0, sticky="w")
        email_entry = tk.Entry(frame)
        email_entry.insert(0, user[3])
        email_entry.grid(row=5, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Mot de passe:").grid(row=6, column=0, sticky="w")
        password_entry = tk.Entry(frame, show="*")
        password_entry.insert(0, user[4])
        password_entry.grid(row=7, column=0, padx=0, pady=0, sticky="ew")

        tk.Label(frame, text="Département:").grid(row=8, column=0, sticky="w")
        department_combobox = ttk.Combobox(frame,
                                           values=["Direction Régionale", "Complexe de Formation Professionnel"],
                                           state="readonly")
        department_combobox.set(user[5])
        department_combobox.grid(row=9, column=0, padx=0, pady=0, sticky="ew")

        tk.Button(frame, text="Mettre à jour l'utilisateur", command=update_user_in_db).grid(row=10, column=0, pady=5, sticky="ew")

    button_frame = ttk.Frame(user_window)
    button_frame.pack(fill='x')

    tk.Button(button_frame, text="Ajouter un utilisateur", command=add_user_window).pack(side='left', padx=5, pady=5)
    tk.Button(button_frame, text="Supprimer un utilisateur", command=delete_user).pack(side='left', padx=5, pady=5)
    tk.Button(button_frame, text="Mettre à jour un utilisateur", command=update_user_window).pack(side='left', padx=5, pady=5)

    load_users()

def initialize_main_application():
    global treeview, imported_file_path, cf_combobox, efp_combobox, vol_combobox, action_entry, objet_entry, priorité_spinbox, statut_combobox, observation_entry, deadline_entry, email_button

    imported_file_path = ""

    frame = ttk.Frame(root, style="Card")
    frame.pack()

    widgets_frame = ttk.LabelFrame(frame, text="Insertion")
    widgets_frame.grid(row=0, column=0)

    cf_options = ["CF AIN AOUDA TAMESNA", "CF HAY NAHDA RABAT", "CF HAY RIAD RABAT", "CF KHEMISSET",
                  "CF MAAMORA KENITRA", "CF SALE I", "CF SALE II", "CF SEBOU SAKNIA KENITRA", "CF SIDI KACEM",
                  "CF TEMARA", "CF YACOUB EL MANSOUR", "INC DAR ESSALAM RABAT", "ISTA SIDI SLIMANE",
                  "CFMHT GUICH LOUDAYA", "CMC RSK"]

    efp_options_dict = {
        "CF AIN AOUDA TAMESNA": ["CFPM SKHIRAT", "CFJ SKHIRAT", "IS BTP & MC TAMESNA", "CFMRA AIN AOUDA",
                                 "ISTA AIN AOUDA"],
        "CF HAY NAHDA RABAT": ["ISTA HAY NAHDA RABAT", "ISTA CONFECTION RABAT", "CENTRE MIXTE HAY NAHDA RABAT"],
        "CF HAY RIAD RABAT": ["ISTA HAY RIAD RABAT", "ISTA NTIC RABAT", "CFMSE YM RABAT",
                              "CROISSANT ROUGE MAROCAIN RABAT"],
        "CF KHEMISSET": ["CSM TIFLET", "CF SIDI ABDERRAZAK", "ISTA KHEMISSET", "CF SIDI ALLAL BAHRAOUI",
                         "ISTA 2 KHEMISSET", "CF ROMMANI"],
        "CF MAAMORA KENITRA": ["ISTA MAAMORA KENITRA", "ISTA INDUSTRIEL MAAMORA KENITRA", "CFPAE OULED OUJIH KENITRA",
                               "CFMA KENITRA", "ISIA KENITRA", "ISMTRL KENITRA"],
        "CF SALE I": ["ISTA HAY SALAM SALE", "CFIJ BETTANA", "CFIJ AL KARIA", "ISTA SALA ALJADIA", "CNMH SALE"],
        "CF SALE II": ["ISTA CHMAOU SALE", "ISMA SALE", "CFIJ BOUKNADEL", "ISTA BOUKNADEL", "CSEDCJ HAY ERRAHMA SALE",
                       "ISTA LAMKINSIA SALE", "CFM SAID HAJJI SALE"],
        "CF SEBOU SAKNIA KENITRA": ["CFIJ SAKNIA KENITRA", "CENTRE FEMME SAKNIA KENITRA", "CFIJ SIDI TAIBI",
                                    "ITA SAKNIA KENITRA", "ITA SEBOU KENITRA", "IS BTP KENITRA", "CFMP SOUK LARBAA"],
        "CF SIDI KACEM": ["ISTA SIDI KACEM", "ITA SIDI KACEM", "ISTA BELKSIRI", "CFP JORF EL MELHA"],
        "CF TEMARA": ["ISTA TEMARA", "CDCJ MASSIRA TEMARA", "ISTA AIN ATIQ", "CFPAE TEMARA"],
        "CF YACOUB EL MANSOUR": ["ITA HR RABAT", "ISTAG & ITAG YM RABAT", "ISTA YM RABAT", "CFPI AL MAJD RABAT",
                                 "CFIJ YM RABAT", "CFMMER RABAT"],
        "INC DAR ESSALAM RABAT": ["INC RABAT"],
        "ISTA SIDI SLIMANE": ["ISTA SIDI SLIMANE"],
        "CFMHT GUICH LOUDAYA": ["CFMHT GUICH LOUDAYA"],
        "CMC RSK": ["CMC RSK"]
    }

    vol_options = ["DGAM", "ARF", "APF"]

    cf_combobox = ttk.Combobox(widgets_frame, values=cf_options, state="readonly")
    cf_combobox.set("CF")
    cf_combobox.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

    efp_combobox = ttk.Combobox(widgets_frame, state="readonly")
    efp_combobox.set("EFP")
    efp_combobox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

    def update_efp_options(event):
        selected_cf = cf_combobox.get()
        efp_options = efp_options_dict.get(selected_cf, [])
        efp_combobox['values'] = efp_options
        if efp_options:
            efp_combobox.set(efp_options[0])
        else:
            efp_combobox.set("Sélectionner EFP")

    cf_combobox.bind("<<ComboboxSelected>>", update_efp_options)

    vol_combobox = ttk.Combobox(widgets_frame, values=vol_options, state="readonly")
    vol_combobox.set("VOL")
    vol_combobox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

    action_entry = ttk.Entry(widgets_frame)
    action_button = ttk.Button(widgets_frame, text="Action",
                               command=lambda: open_text_editor("ACTION", action_entry))
    action_button.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

    objet_entry = ttk.Entry(widgets_frame)
    objet_button = ttk.Button(widgets_frame, text="Objet", command=lambda: open_text_editor("OBJET", objet_entry))
    objet_button.grid(row=4, column=0, padx=5, pady=5, sticky="ew")

    cf_efp_entry = ttk.Entry(widgets_frame)
    cf_efp_button = ttk.Button(widgets_frame, text="CFP/EFP", command=lambda: open_text_editor("CFP/EFP", cf_efp_entry))
    cf_efp_button.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

    priorities = ["Basse", "Moyenne", "Haute"]
    priorité_spinbox = ttk.Spinbox(widgets_frame, values=priorities)
    priorité_spinbox.insert(0, "Priorité")
    priorité_spinbox.grid(row=6, column=0, padx=5, pady=5, sticky="ew")

    statut_combobox = ttk.Combobox(widgets_frame, values=["Non démarrée", "En cours", "Terminé"], state="readonly")
    statut_combobox.set("Statut")
    statut_combobox.grid(row=7, column=0, padx=5, pady=5, sticky="ew")

    deadline_entry = ttk.Combobox(widgets_frame, width=20, textvariable=tk.StringVar(), state="readonly")
    deadline_entry.set("Date de deadline")
    cal_deadline = Calendar(widgets_frame, selectmode="day", year=2024, month=3, day=1)
    deadline_entry.grid(row=8, column=0, padx=5, pady=5, sticky="ew")
    deadline_entry.bind("<FocusIn>", lambda event: toggle_calendar(cal_deadline))
    deadline_entry.bind("<Button-1>", lambda event: toggle_calendar(cal_deadline))
    cal_deadline.bind("<<CalendarSelected>>", lambda event: select_date(cal_deadline, deadline_entry))

    observation_entry = ttk.Entry(widgets_frame)
    observation_button = ttk.Button(widgets_frame, text="Observation",
                                    command=lambda: open_text_editor("OBSERVATIONS", observation_entry))
    observation_button.grid(row=9, column=0, padx=5, pady=5, sticky="ew")

    import_button = ttk.Button(widgets_frame, text="Importer", command=imported_file)
    import_button.grid(row=10, column=0, padx=5, pady=5, sticky="ew")

    separator = ttk.Separator(widgets_frame, orient="horizontal")
    separator.grid(row=11, column=0, padx=5, pady=5, sticky="ew")

    button = ttk.Button(widgets_frame, text="Insérer", style="Accent.TButton", command=insert_row)
    button.grid(row=12, column=0, padx=5, pady=5, sticky="nsew")

    delete_button = ttk.Button(widgets_frame, text="Supprimer", command=delete_row)
    delete_button.grid(row=13, column=0, padx=5, pady=5, sticky="ew")

    email_button = ttk.Button(widgets_frame, text="Email de rappel", command=generate_email)
    email_button.grid(row=14, column=0, padx=5, pady=5, sticky="ew")

    manage_db_button = ttk.Button(widgets_frame, text="Gérer la base de données", command=manage_database)
    manage_db_button.grid(row=15, column=0, padx=5, pady=5, sticky="ew")

    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)

    # Add vertical and horizontal scrollbars to the treeview
    treeScrollY = ttk.Scrollbar(treeFrame, orient="vertical")
    treeScrollY.pack(side="right", fill="y")

    treeScrollX = ttk.Scrollbar(treeFrame, orient="horizontal")
    treeScrollX.pack(side="bottom", fill="x")

    global cols
    cols = new_cols[3:]  # Only include columns starting from "ACTION"

    # Configure Treeview style
    style = ttk.Style()
    style.configure("Treeview", font=("Arial", 9), rowheight=30)  # Set the font size and row height

    treeview = ttk.Treeview(treeFrame, show="headings", columns=cols, height=20, selectmode="extended",
                            style="Treeview")
    treeview.bind("<ButtonRelease-1>", show_detailed_information)

    for col_index, col_name in enumerate(cols):
        treeview.heading(col_index, text=col_name)
        treeview.column(col_index, anchor="center")

    treeview.pack(expand=True, fill="both")
    treeScrollY.config(command=treeview.yview)
    treeScrollX.config(command=treeview.xview)
    treeview.configure(yscrollcommand=treeScrollY.set, xscrollcommand=treeScrollX.set)

    detailed_info_frame = ttk.LabelFrame(frame, text="Détail", height=500)
    detailed_info_frame.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")

    detail_buttons = {
        "ACTION": ttk.Button(detailed_info_frame, text="ACTION", command=lambda: edit_detail("ACTION"), style="Accent.TButton"),
        "OBJET": ttk.Button(detailed_info_frame, text="OBJET", command=lambda: edit_detail("OBJET"), style="Accent.TButton"),
        "CFP/ EFP": ttk.Button(detailed_info_frame, text="CFP/ EFP", command=lambda: edit_detail("CFP/ EFP"), style="Accent.TButton"),
        "PRIORITÉ": ttk.Button(detailed_info_frame, text="PRIORITÉ", command=lambda: edit_detail("PRIORITÉ"), style="Accent.TButton"),
        "STATUT": ttk.Button(detailed_info_frame, text="STATUT", command=lambda: edit_detail("STATUT"), style="Accent.TButton"),
        "DEADLINE": ttk.Button(detailed_info_frame, text="DEADLINE", command=lambda: edit_detail("DEADLINE"), style="Accent.TButton"),
        "OBSERVATIONS": ttk.Button(detailed_info_frame, text="OBSERVATIONS", command=lambda: edit_detail("OBSERVATIONS"), style="Accent.TButton"),
        "RESSOURCES": ttk.Button(detailed_info_frame, text="RESSOURCES", command=lambda: edit_detail("RESSOURCES"), style="Accent.TButton"),
    }

    for i, (col_name, button) in enumerate(detail_buttons.items()):
        button.grid(row=0, column=i, padx=5, pady=5, sticky="ew")

    load_data()

def open_text_editor(field_name, entry_widget):
    def save_text():
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, text_editor.get("1.0", tk.END).strip())
        editor_window.destroy()

    editor_window = tk.Toplevel(root)
    editor_window.title(f"Modifier {field_name}")
    apply_theme(editor_window)

    text_editor = tk.Text(editor_window, wrap='word', height=20, width=80)
    text_editor.pack(pady=10, padx=10)

    text_editor.insert(tk.END, entry_widget.get())

    save_button = ttk.Button(editor_window, text="Sauvegarder", command=save_text)
    save_button.pack(pady=5)

    editor_window.transient(root)
    editor_window.grab_set()
    root.wait_window(editor_window)

def edit_detail(column_name):
    selected_item = treeview.selection()[0]
    column_index = cols.index(column_name)
    current_value = treeview.item(selected_item, "values")[column_index]

    def save_text():
        new_value = text_editor.get("1.0", tk.END).strip()
        treeview.set(selected_item, column=column_name, value=new_value)
        editor_window.destroy()

    editor_window = tk.Toplevel(root)
    editor_window.title(f"Modifier {column_name}")
    apply_theme(editor_window)

    text_editor = tk.Text(editor_window, wrap='word', height=20, width=80)
    text_editor.pack(pady=10, padx=10)
    text_editor.insert(tk.END, current_value)

    save_button = ttk.Button(editor_window, text="Sauvegarder", command=save_text)
    save_button.pack(pady=5)

    editor_window.transient(root)
    editor_window.grab_set()
    root.wait_window(editor_window)

def toggle_calendar(cal, event=None):
    cal.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

def select_date(cal, combobox):
    selected_date = cal.get_date()
    combobox.set(selected_date)
    cal.grid_remove()

def show_detailed_information(event):
    selected_item = treeview.selection()[0]
    for col_name, button in detail_buttons.items():
        column_index = cols.index(col_name)
        button.config(text=f"Modifier {col_name}: {treeview.item(selected_item, 'values')[column_index]}")

if __name__ == "__main__":
    root = tk.Tk()
    root.title("OFPPT Suivi d'Actions")
    root.withdraw()
    create_database()

    logo_path = r'C:\Users\Nasser\Downloads\logoprogramme.png'
    logo = PhotoImage(file=logo_path)
    root.logo = logo

    root.iconphoto(False, logo)

    root.deiconify()

    current_user = None

    show_login_window()
    root.mainloop()
